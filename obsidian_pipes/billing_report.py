"""
title: n8n Billing Query Report
author: Bishr
version: 0.6
description: Triggers n8n billing workflow with required parameters, date shortcuts, batch support, and Excel upload
requirements: openpyxl
"""

from pydantic import BaseModel, Field
from typing import Optional, Union, Generator, Iterator
import requests
import json
import urllib3
import base64
import io
from datetime import datetime, timedelta
from calendar import monthrange
from pathlib import Path

# Suppress SSL warnings for self-signed certificates
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Open WebUI imports for file access
try:
    from open_webui.models.files import Files
    from open_webui.storage.provider import Storage
    OPENWEBUI_AVAILABLE = True
except ImportError:
    OPENWEBUI_AVAILABLE = False
    print("[Billing Pipe] Warning: Open WebUI modules not available - file upload may not work")

# Module-level tracking to prevent duplicate processing across all instances
# This survives class re-instantiation
_PROCESSED_MESSAGES = set()
_PROCESSING_LOCK = set()  # Track messages currently being processed
_N8N_CALL_COUNT = 0  # Debug counter


class Pipe:
    class Valves(BaseModel):
        """Configurable settings - accessible from the Open WebUI admin UI"""

        N8N_WEBHOOK_URL: str = Field(
            default="https://10.1.21.253:5678/webhook/211faa0e-2b99-4ab4-b270-b27992604384",
            description="Your n8n webhook URL",
        )
        DEFAULT_GROUP_FACILITIES: str = Field(
            default="false", 
            description="Default group_facilities value (true/false)"
        )

    def __init__(self):
        self.type = "pipe"
        self.id = "n8n_billing_report"
        self.name = "n8n Billing Report"
        self.valves = self.Valves()
        # Enable file handling
        self.file_handler = True

    def pipe(
        self,
        body: dict,
        __user__: Optional[dict] = None,
        __event_emitter__=None,
        __files__: Optional[list] = None,
    ) -> Union[str, Generator, Iterator]:
        """
        Main pipe function - called when a user sends a message.
        
        Supports:
        - Simple commands: "run report"
        - Parameter overrides: "protocol=MK9999999"
        - Date shortcuts: "this_month", "last_month", "q1", etc.
        - Batch processing: multiple protocols comma-separated or one per line
        - Excel file upload: Upload a file with columns Protocol, start_date, end_date, group_facilities
        """
        global _PROCESSED_MESSAGES, _PROCESSING_LOCK
        
        # Prevent duplicate processing - Open WebUI may call pipe multiple times
        metadata = body.get("metadata", {})
        message_id = metadata.get("message_id", "")
        
        if message_id:
            # Check if already processed or currently processing
            if message_id in _PROCESSED_MESSAGES:
                print(f"[Billing Pipe] Skipping already-processed request: {message_id}")
                return ""
            
            if message_id in _PROCESSING_LOCK:
                print(f"[Billing Pipe] Skipping parallel request (already in progress): {message_id}")
                return ""
            
            # Mark as currently processing
            _PROCESSING_LOCK.add(message_id)
            print(f"[Billing Pipe] Processing new request: {message_id}")
        
        try:
            result = self._execute_pipe(body, __user__, __event_emitter__, __files__, metadata)
            
            # Mark as completed
            if message_id:
                _PROCESSED_MESSAGES.add(message_id)
                _PROCESSING_LOCK.discard(message_id)
                # Keep set from growing too large
                if len(_PROCESSED_MESSAGES) > 100:
                    _PROCESSED_MESSAGES = set(list(_PROCESSED_MESSAGES)[-50:])
            
            return result
        except Exception as e:
            # Remove from processing lock on error
            if message_id:
                _PROCESSING_LOCK.discard(message_id)
            raise e
    
    def _execute_pipe(
        self,
        body: dict,
        __user__: Optional[dict],
        __event_emitter__,
        __files__: Optional[list],
        metadata: dict,
    ) -> Union[str, Generator, Iterator]:
        """Actual pipe execution logic."""
        
        # Get the user's message
        messages = body.get("messages", [])
        user_message = ""
        if messages:
            last_message = messages[-1]
            if last_message.get("role") == "user":
                user_message = last_message.get("content", "")
        
        # Check for files in various locations
        files = __files__ or []
        
        # Also check metadata for files (metadata already defined above)
        if not files and metadata.get("files"):
            files = metadata.get("files", [])
        
        # Also check body for files directly
        if not files and body.get("files"):
            files = body.get("files", [])
        
        # Debug: Log file detection info (visible in Open WebUI logs)
        if files:
            print(f"[Billing Pipe] Files detected: {len(files)}")
            for i, f in enumerate(files):
                if isinstance(f, dict):
                    fname = f.get("name", f.get("filename", "unknown"))
                    fkeys = list(f.keys())
                    print(f"[Billing Pipe] File {i+1}: {fname}, keys: {fkeys}")
        else:
            print("[Billing Pipe] No files detected in request")
        
        # Check if there's an Excel/CSV file uploaded
        excel_batch_items = self._parse_excel_files(files)
        if excel_batch_items:
            return self._process_batch(excel_batch_items, {}, __user__, source="📁 File upload")
        
        # Parse the message for parameters and batch items
        params, batch_items = self._parse_message(user_message)
        
        # If batch items found, process multiple
        if batch_items:
            return self._process_batch(batch_items, params, __user__, source="📝 Text input")
        
        # Check if user wants help
        if "help" in user_message.lower():
            return self._show_help()
        
        # Check if user provided meaningful parameters
        has_protocol = "protocol" in params
        has_dates = "start_date" in params or "end_date" in params
        has_run_command = any(cmd in user_message.lower() for cmd in ["run", "generate", "create", "execute", "go"])
        
        # If no parameters provided, show help instead of running default
        if not has_protocol and not has_dates and not has_run_command:
            return self._show_help()
        
        # Single request with provided parameters
        return self._process_single(params, __user__)

    def _show_help(self) -> str:
        """Show help message when no parameters are provided."""
        return """## 📊 Billing Report Generator

Welcome! I can generate billing reports from your n8n workflow.

### Quick Start

**Single Report:**
```
protocol=MK3543006 start_date=2025-01-01 end_date=2025-03-31
```

**With Group Facilities:**
```
protocol=MK3543006 start_date=2025-01-01 end_date=2025-03-31 group_facilities=true
```

### Date Shortcuts

| Shortcut | Description |
|----------|-------------|
| `this_month` | Current month |
| `last_month` | Previous month |
| `this_year` | Current year |
| `q1`, `q2`, `q3`, `q4` | Quarterly |

**Example:** `protocol=MK3543006 this_month`

### Batch Processing

**Multiple protocols (comma-separated):**
```
protocols=MK3543006,MK3543007,MK1084004 q1
```

**Upload Excel/CSV file:**
Upload a file with columns: `Protocol`, `start_date`, `end_date`, `group_facilities`

---

💡 *Type your parameters above or upload a file to get started!*
"""

    def _parse_excel_files(self, files: list) -> list:
        """
        Parse Excel/CSV files for batch processing.
        
        Expected columns: Protocol, start_date, end_date, group_facilities
        (Column names are case-insensitive)
        """
        batch_items = []
        
        if not files:
            return batch_items
        
        for file_info in files:
            # Get file data - handle different file info formats
            file_data = None
            file_name = ""
            is_excel = False
            is_csv = False
            
            if isinstance(file_info, dict):
                file_name = file_info.get("name", "") or file_info.get("filename", "")
                file_id = file_info.get("id", "")
                
                print(f"[Billing Pipe] Processing file: name={file_name}, id={file_id}")
                
                # Check if it's an Excel or CSV file
                is_excel = any(file_name.lower().endswith(ext) for ext in [".xlsx", ".xls"])
                is_csv = file_name.lower().endswith(".csv")
                
                if not is_excel and not is_csv:
                    print(f"[Billing Pipe] Skipping non-Excel/CSV file: {file_name}")
                    continue
                
                # Method 1: Get file from Open WebUI storage by ID (primary method)
                if file_id and OPENWEBUI_AVAILABLE:
                    try:
                        file_record = Files.get_file_by_id(file_id)
                        if file_record and file_record.path:
                            storage_path = Storage.get_file(file_record.path)
                            storage_path = Path(storage_path)
                            if storage_path.is_file():
                                with open(storage_path, "rb") as f:
                                    file_data = f.read()
                                print(f"[Billing Pipe] Read file from storage: {storage_path}")
                    except Exception as e:
                        print(f"[Billing Pipe] Error reading from storage: {e}")
                
                # Method 2: Base64 encoded data (fallback)
                if not file_data and "data" in file_info:
                    file_data = file_info["data"]
                    if isinstance(file_data, str):
                        # Remove data URL prefix if present
                        if "," in file_data:
                            file_data = file_data.split(",", 1)[1]
                        file_data = base64.b64decode(file_data)
                    print(f"[Billing Pipe] Read file from base64 data")
                
                # Method 3: Direct content (fallback)
                if not file_data and "content" in file_info:
                    file_data = file_info["content"]
                    if isinstance(file_data, str):
                        file_data = base64.b64decode(file_data)
                    print(f"[Billing Pipe] Read file from content field")
                
                # Method 4: File path (fallback)
                if not file_data and "path" in file_info:
                    try:
                        with open(file_info["path"], "rb") as f:
                            file_data = f.read()
                        print(f"[Billing Pipe] Read file from path: {file_info['path']}")
                    except Exception as e:
                        print(f"[Billing Pipe] Error reading from path: {e}")
                
                # Method 5: URL download (fallback)
                if not file_data and "url" in file_info:
                    try:
                        response = requests.get(file_info["url"], timeout=30)
                        file_data = response.content
                        print(f"[Billing Pipe] Downloaded file from URL")
                    except Exception as e:
                        print(f"[Billing Pipe] Error downloading from URL: {e}")
            
            if not file_data:
                print(f"[Billing Pipe] No file data found for: {file_name}")
                continue
            
            print(f"[Billing Pipe] Successfully loaded file: {file_name} ({len(file_data)} bytes)")
            
            # Parse CSV files
            if is_csv:
                try:
                    import csv
                    content = file_data.decode('utf-8')
                    reader = csv.DictReader(io.StringIO(content))
                    
                    for row in reader:
                        # Normalize column names
                        normalized_row = {k.lower().strip(): v for k, v in row.items() if k}
                        
                        protocol = None
                        for key in normalized_row:
                            if "protocol" in key:
                                protocol = normalized_row[key]
                                break
                        
                        if not protocol:
                            continue
                        
                        item = {"protocol": str(protocol).strip()}
                        
                        for key, value in normalized_row.items():
                            if "start" in key and "date" in key:
                                item["start_date"] = self._normalize_date(value)
                            elif "end" in key and "date" in key:
                                item["end_date"] = self._normalize_date(value)
                            elif "group" in key or "facilities" in key:
                                item["group_facilities"] = str(value).lower().strip()
                        
                        batch_items.append(item)
                    
                    continue  # Move to next file
                    
                except Exception as e:
                    print(f"Error parsing CSV file {file_name}: {e}")
                    continue
            
            # Parse Excel files
            try:
                from openpyxl import load_workbook
            except ImportError:
                print("openpyxl not installed, skipping Excel file")
                continue
                
            try:
                workbook = load_workbook(filename=io.BytesIO(file_data), read_only=True)
                sheet = workbook.active
                
                # Get headers from first row
                headers = []
                header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = [str(h).lower().strip() if h else "" for h in header_row]
                
                # Map column names to indices
                column_map = {}
                for idx, header in enumerate(headers):
                    # Normalize column names
                    if "protocol" in header:
                        column_map["protocol"] = idx
                    elif "start" in header and "date" in header:
                        column_map["start_date"] = idx
                    elif "end" in header and "date" in header:
                        column_map["end_date"] = idx
                    elif "group" in header or "facilities" in header:
                        column_map["group_facilities"] = idx
                
                # Check if we have at least protocol column
                if "protocol" not in column_map:
                    continue
                
                # Parse data rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    protocol = row[column_map["protocol"]] if "protocol" in column_map else None
                    
                    if not protocol:
                        continue  # Skip empty rows
                    
                    item = {"protocol": str(protocol).strip()}
                    
                    if "start_date" in column_map and row[column_map["start_date"]]:
                        start_val = row[column_map["start_date"]]
                        item["start_date"] = self._normalize_date(start_val)
                    
                    if "end_date" in column_map and row[column_map["end_date"]]:
                        end_val = row[column_map["end_date"]]
                        item["end_date"] = self._normalize_date(end_val)
                    
                    if "group_facilities" in column_map and row[column_map["group_facilities"]] is not None:
                        gf_val = row[column_map["group_facilities"]]
                        item["group_facilities"] = str(gf_val).lower().strip()
                    
                    batch_items.append(item)
                
                workbook.close()
                
            except Exception as e:
                # Log error but continue
                print(f"Error parsing Excel file {file_name}: {e}")
                continue
        
        # Deduplicate batch items by protocol (in case of duplicate rows or files)
        seen_protocols = set()
        unique_items = []
        for item in batch_items:
            protocol = item.get("protocol", "")
            if protocol and protocol not in seen_protocols:
                seen_protocols.add(protocol)
                unique_items.append(item)
        
        print(f"[Billing Pipe] Parsed {len(unique_items)} unique protocols from files (original: {len(batch_items)})")
        return unique_items

    def _normalize_date(self, date_val) -> str:
        """Convert various date formats to YYYY-MM-DD string."""
        if date_val is None:
            return ""
        
        # If it's already a datetime object
        if isinstance(date_val, datetime):
            return date_val.strftime("%Y-%m-%d")
        
        # If it's a string, try to parse common formats
        date_str = str(date_val).strip()
        
        # Already in correct format
        if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
            return date_str
        
        # Try common formats
        formats = [
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%m-%d-%Y",
            "%d-%m-%Y",
        ]
        
        for fmt in formats:
            try:
                dt = datetime.strptime(date_str[:10], fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
        
        # Return as-is if we can't parse
        return date_str

    def _parse_message(self, message: str) -> tuple[dict, list]:
        """
        Parse message for parameters and detect batch input.
        
        Supports:
        - key=value pairs
        - Date shortcuts (this_month, last_month, this_year, q1-q4)
        - Multiple protocols (comma-separated or newline-separated)
        """
        params = {}
        batch_items = []
        
        message_lower = message.lower()
        now = datetime.now()
        
        # ===== DATE SHORTCUTS =====
        
        if "this_month" in message_lower or "this month" in message_lower:
            start = now.replace(day=1)
            _, last_day = monthrange(now.year, now.month)
            end = now.replace(day=last_day)
            params["start_date"] = start.strftime("%Y-%m-%d")
            params["end_date"] = end.strftime("%Y-%m-%d")
        
        elif "last_month" in message_lower or "last month" in message_lower:
            first_of_this_month = now.replace(day=1)
            last_month_end = first_of_this_month - timedelta(days=1)
            last_month_start = last_month_end.replace(day=1)
            params["start_date"] = last_month_start.strftime("%Y-%m-%d")
            params["end_date"] = last_month_end.strftime("%Y-%m-%d")
        
        elif "this_year" in message_lower or "this year" in message_lower:
            params["start_date"] = f"{now.year}-01-01"
            params["end_date"] = f"{now.year}-12-31"
        
        elif "last_year" in message_lower or "last year" in message_lower:
            params["start_date"] = f"{now.year - 1}-01-01"
            params["end_date"] = f"{now.year - 1}-12-31"
        
        # Quarter shortcuts
        quarters = {
            "q1": (1, 3), 
            "q2": (4, 6), 
            "q3": (7, 9), 
            "q4": (10, 12)
        }
        for q, (start_m, end_m) in quarters.items():
            if q in message_lower:
                _, last_day = monthrange(now.year, end_m)
                params["start_date"] = f"{now.year}-{start_m:02d}-01"
                params["end_date"] = f"{now.year}-{end_m:02d}-{last_day}"
                break
        
        # ===== KEY=VALUE PARSING =====
        
        parts = message.split()
        for part in parts:
            if "=" in part:
                key, value = part.split("=", 1)
                key = key.strip().lower()
                value = value.strip()
                params[key] = value
        
        # ===== BATCH PROCESSING =====
        
        # Check for batch: multiple protocols (comma-separated)
        if "protocols" in params:
            protocols = [
                p.strip() 
                for p in params["protocols"].replace(",", "\n").split("\n") 
                if p.strip()
            ]
            if len(protocols) > 1:
                for protocol in protocols:
                    batch_items.append({**params, "protocol": protocol})
                del params["protocols"]
        
        # Check for newline-separated protocol list in message
        lines = message.strip().split("\n")
        if len(lines) > 1:
            potential_protocols = []
            for line in lines:
                line = line.strip()
                # Skip empty lines, key=value pairs, markdown headers, lines with colons, and long lines
                if not line or "=" in line or ":" in line or line.startswith("#") or len(line) > 30:
                    continue
                # Check if it looks like a protocol ID (starts with letters, contains numbers)
                # Protocol IDs typically look like: MK3543006, ABC123, etc.
                if line[0].isalpha() and any(c.isdigit() for c in line):
                    potential_protocols.append(line)
            
            if len(potential_protocols) > 1:
                for protocol in potential_protocols:
                    batch_items.append({**params, "protocol": protocol})
        
        return params, batch_items

    def _format_date(self, date_str: str, is_end: bool = False) -> str:
        """
        Add time component to date string if not present.
        
        - Start dates get 00:00:00
        - End dates get 23:59:59
        """
        if not date_str:
            return date_str
        
        # If already has time component, return as-is
        if " " in date_str or "T" in date_str:
            return date_str
        
        # Add appropriate time
        if is_end:
            return f"{date_str} 23:59:59"
        else:
            return f"{date_str} 00:00:00"

    def _build_payload(self, params: dict, __user__: Optional[dict]) -> dict:
        """Build the payload for n8n webhook."""
        start_date = params.get("start_date", "")
        end_date = params.get("end_date", "")
        
        return {
            "user_id": __user__.get("id") if __user__ else None,
            "user_email": __user__.get("email") if __user__ else None,
            "user_name": __user__.get("name") if __user__ else None,
            "form_data": {
                "protocol": params.get("protocol", ""),
                "start_date": self._format_date(start_date, is_end=False),
                "end_date": self._format_date(end_date, is_end=True),
                "group_facilities": params.get(
                    "group_facilities", self.valves.DEFAULT_GROUP_FACILITIES
                ),
            },
        }
    
    def _validate_params(self, params: dict) -> tuple[bool, str]:
        """Validate that required parameters are provided."""
        missing = []
        
        if not params.get("protocol"):
            missing.append("protocol")
        if not params.get("start_date"):
            missing.append("start_date")
        if not params.get("end_date"):
            missing.append("end_date")
        
        if missing:
            return False, f"Missing required parameters: **{', '.join(missing)}**"
        
        return True, ""

    def _call_n8n(self, payload: dict) -> dict:
        """Call the n8n webhook and return the response."""
        global _N8N_CALL_COUNT
        
        # Log the call for debugging
        form_data = payload.get("form_data", {})
        payload_key = f"{form_data.get('protocol')}|{form_data.get('start_date')}|{form_data.get('end_date')}"
        
        _N8N_CALL_COUNT += 1
        print(f"[Billing Pipe] Making n8n call #{_N8N_CALL_COUNT}: {payload_key}")
        
        response = requests.post(
            self.valves.N8N_WEBHOOK_URL,
            json=payload,
            headers={"Content-Type": "application/json"},
            timeout=300,  # 5 minutes for long-running reports
            verify=False,  # Allow self-signed certificates
        )
        response.raise_for_status()
        
        try:
            return response.json()
        except json.JSONDecodeError:
            return {"message": response.text}

    def _process_single(self, params: dict, __user__: Optional[dict]) -> str:
        """Process a single report request."""
        # Validate required parameters
        is_valid, error_msg = self._validate_params(params)
        if not is_valid:
            return f"""## ⚠️ Missing Parameters

{error_msg}

**Required format:**
```
protocol=MK3543006 start_date=2025-01-01 end_date=2025-03-31
```

Or use a date shortcut:
```
protocol=MK3543006 this_month
```

💡 *Type "help" to see all options.*
"""
        
        payload = self._build_payload(params, __user__)
        
        try:
            result = self._call_n8n(payload)
            return self._format_response(result, payload["form_data"])
        except requests.exceptions.Timeout:
            return "⚠️ **Error**: The workflow timed out. Please try again."
        except requests.exceptions.RequestException as e:
            return f"⚠️ **Error**: Failed to connect to n8n workflow.\n\n```\n{str(e)}\n```"
        except Exception as e:
            return f"⚠️ **Error**: {str(e)}"

    def _process_batch(
        self, batch_items: list, base_params: dict, __user__: Optional[dict], source: str = "text"
    ) -> str:
        """Process multiple report requests."""
        global _N8N_CALL_COUNT
        
        print(f"[Billing Pipe] _process_batch called with {len(batch_items)} items, source={source}")
        print(f"[Billing Pipe] Current n8n call count before batch: {_N8N_CALL_COUNT}")
        
        output = f"## 📊 Batch Report Processing\n\n"
        output += f"*Source: {source}*\n\n"
        output += f"Processing **{len(batch_items)}** reports...\n\n"
        
        results = []
        for i, item_params in enumerate(batch_items, 1):
            # Merge base params with item-specific params
            merged_params = {**base_params, **item_params}
            payload = self._build_payload(merged_params, __user__)
            protocol = payload["form_data"]["protocol"]
            start_date = payload["form_data"]["start_date"]
            end_date = payload["form_data"]["end_date"]
            
            output += f"### Report {i}: `{protocol}`\n"
            output += f"*{start_date} → {end_date}*\n\n"
            
            try:
                result = self._call_n8n(payload)
                message = result.get("message", "Completed")
                output += f"✅ {message}\n\n"
                results.append({
                    "protocol": protocol, 
                    "success": True, 
                    "result": result
                })
            except Exception as e:
                output += f"❌ Failed: {str(e)}\n\n"
                results.append({
                    "protocol": protocol, 
                    "success": False, 
                    "error": str(e)
                })
        
        # Summary
        success_count = sum(1 for r in results if r["success"])
        output += f"---\n\n"
        output += f"**Summary:** {success_count}/{len(batch_items)} reports generated successfully.\n"
        
        return output

    def _format_response(self, result: dict, form_data: dict) -> str:
        """Format the n8n response for display in chat."""
        output = "## ✅ Query Report Executed\n\n"
        output += "### Parameters Used:\n"
        output += f"- **Protocol**: `{form_data['protocol']}`\n"
        output += f"- **Start Date**: `{form_data['start_date']}`\n"
        output += f"- **End Date**: `{form_data['end_date']}`\n"
        output += f"- **Group Facilities**: `{form_data['group_facilities']}`\n\n"
        
        # Handle n8n response
        if isinstance(result, dict):
            message = (
                result.get("message") 
                or result.get("text") 
                or result.get("response")
            )
            if message:
                output += f"### Response:\n{message}\n\n"
            
            file_url = (
                result.get("file_url")
                or result.get("fileUrl")
                or result.get("download_url")
                or result.get("downloadUrl")
                or result.get("url")
            )
            file_name = (
                result.get("file_name") 
                or result.get("fileName") 
                or result.get("filename")
                or "Download"
            )
            
            if file_url:
                output += f"### Download:\n[📥 {file_name}]({file_url})\n\n"
            
            # Handle data tables
            data = result.get("data")
            if data and isinstance(data, list) and len(data) > 0:
                output += "### Data:\n"
                if isinstance(data[0], dict):
                    headers = list(data[0].keys())
                    output += "| " + " | ".join(headers) + " |\n"
                    output += "| " + " | ".join(["---"] * len(headers)) + " |\n"
                    for row in data[:20]:  # Limit to 20 rows
                        output += (
                            "| "
                            + " | ".join(str(row.get(h, "")) for h in headers)
                            + " |\n"
                        )
                    if len(data) > 20:
                        output += f"\n*...and {len(data) - 20} more rows*\n"
        elif isinstance(result, list):
            output += f"### Raw Response:\n```json\n{json.dumps(result, indent=2)}\n```\n"
        else:
            output += f"### Response:\n{result}\n"
        
        return output
